"""Directory processing routes."""

from __future__ import annotations

import asyncio
import logging
import os
from datetime import datetime
from uuid import uuid4

from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

from fastapi import APIRouter, HTTPException
from pydantic import BaseModel, Field
import torch

from ..core.config import settings
from ..services.image_recognition import get_recognition_service
from ..services.translation import get_translation_service
from ..utils.file_handler import (
    build_unique_output_path,
    build_category_output_path,
    build_unique_path_for_name,
    copy_to_output,
    ensure_directory,
    get_file_extension,
    iter_image_files,
    get_camera_make,
    classify_photo_type,
)

logger = logging.getLogger(__name__)
router = APIRouter(prefix="/api/v1", tags=["processing"])

JOB_STORE: dict[str, dict] = {}


def _format_duration(total_seconds: float) -> str:
    """格式化时长为易读格式。"""
    if total_seconds < 60:
        return f"{total_seconds:.1f} 秒"
    elif total_seconds < 3600:
        minutes = total_seconds / 60
        return f"{minutes:.1f} 分钟"
    else:
        hours = total_seconds / 3600
        return f"{hours:.1f} 小时"


def _generate_score_excel(results: list[dict], errors: list[dict], output_path: str) -> None:
    """生成照片评分Excel表格。"""
    wb = Workbook()
    
    # 创建数据工作表
    ws = wb.active
    ws.title = "照片评分表"
    
    # 设置表头样式
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center")
    thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), 
                        top=Side(style='thin'), bottom=Side(style='thin'))
    
    # 表头列
    headers = [
        "序号", "原文件名", "新文件名", "照片类型", 
        "标签", "最高置信度(%)", "评分", "处理状态"
    ]
    
    # 写入表头
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = thin_border
    
    # 设置数据行样式
    data_alignment = Alignment(horizontal="left", vertical="center")
    center_alignment = Alignment(horizontal="center", vertical="center")
    
    # 写入数据
    row_num = 2
    total_score = 0.0
    success_count = 0
    failed_count = len(errors)
    
    for result in results:
        # 提取信息
        original_name = result.get("original_filename", "")
        renamed_name = result.get("renamed_filename", "")
        status = result.get("status", "")
        labels = result.get("labels", [])
        
        # 获取照片类型（从标签中提取）
        photo_type = ""
        max_confidence = 0.0
        
        for label in labels:
            confidence = label.get("confidence_percentage", 0)
            if confidence > max_confidence:
                max_confidence = confidence
        
        # 生成标签字符串
        label_str = ", ".join([label.get("label", "") for label in labels])
        
        # 计算评分（基于最高置信度）
        # 评分规则：置信度 >= 90%: 90-100分, 70-89%: 70-89分, 50-69%: 50-69分, <50%: 0-49分
        if max_confidence >= 90:
            score = round(90 + (max_confidence - 90) * 1)
        elif max_confidence >= 70:
            score = round(70 + (max_confidence - 70) * 0.95)
        elif max_confidence >= 50:
            score = round(50 + (max_confidence - 50) * 0.95)
        else:
            score = round(max_confidence * 0.98)
        
        score = max(0, min(100, score))
        total_score += score
        success_count += 1
        
        # 处理状态映射
        status_display = {
            "renamed": "已重命名",
            "kept_original_name": "保留原名",
        }.get(status, status)
        
        # 写入数据行
        ws.cell(row=row_num, column=1, value=row_num - 1).alignment = center_alignment
        ws.cell(row=row_num, column=2, value=original_name).alignment = data_alignment
        ws.cell(row=row_num, column=3, value=renamed_name).alignment = data_alignment
        ws.cell(row=row_num, column=4, value=photo_type).alignment = data_alignment
        ws.cell(row=row_num, column=5, value=label_str).alignment = data_alignment
        ws.cell(row=row_num, column=6, value=round(max_confidence, 2)).alignment = center_alignment
        ws.cell(row=row_num, column=7, value=round(score, 1)).alignment = center_alignment
        ws.cell(row=row_num, column=8, value=status_display).alignment = center_alignment
        
        row_num += 1
    
    # 写入错误记录
    for error in errors:
        original_name = error.get("filename", "")
        error_msg = error.get("error", "")
        
        ws.cell(row=row_num, column=1, value=row_num - 1).alignment = center_alignment
        ws.cell(row=row_num, column=2, value=original_name).alignment = data_alignment
        ws.cell(row=row_num, column=3, value="").alignment = data_alignment
        ws.cell(row=row_num, column=4, value="").alignment = data_alignment
        ws.cell(row=row_num, column=5, value="").alignment = data_alignment
        ws.cell(row=row_num, column=6, value="").alignment = center_alignment
        ws.cell(row=row_num, column=7, value="").alignment = center_alignment
        ws.cell(row=row_num, column=8, value=f"处理失败: {error_msg}").alignment = data_alignment
        
        row_num += 1
    
    # 添加统计信息
    row_num += 2
    ws.cell(row=row_num, column=1, value="统计信息").font = Font(bold=True)
    
    row_num += 1
    ws.cell(row=row_num, column=2, value="总照片数:").alignment = data_alignment
    ws.cell(row=row_num, column=3, value=len(results) + len(errors)).alignment = center_alignment
    
    row_num += 1
    ws.cell(row=row_num, column=2, value="成功处理:").alignment = data_alignment
    ws.cell(row=row_num, column=3, value=success_count).alignment = center_alignment
    
    row_num += 1
    ws.cell(row=row_num, column=2, value="处理失败:").alignment = data_alignment
    ws.cell(row=row_num, column=3, value=failed_count).alignment = center_alignment
    
    row_num += 1
    ws.cell(row=row_num, column=2, value="成功率:").alignment = data_alignment
    success_rate = (success_count / (len(results) + len(errors))) * 100 if (len(results) + len(errors)) > 0 else 0
    ws.cell(row=row_num, column=3, value=f"{success_rate:.1f}%").alignment = center_alignment
    
    row_num += 1
    ws.cell(row=row_num, column=2, value="平均评分:").alignment = data_alignment
    avg_score = total_score / success_count if success_count > 0 else 0
    ws.cell(row=row_num, column=3, value=f"{avg_score:.1f}").alignment = center_alignment
    
    # 自动调整列宽
    for col in range(1, len(headers) + 1):
        max_length = 0
        column = ws[get_column_letter(col)]
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 50)
        ws.column_dimensions[get_column_letter(col)].width = adjusted_width
    
    # 确保输出目录存在
    output_dir = os.path.dirname(output_path)
    if output_dir and not os.path.exists(output_dir):
        os.makedirs(output_dir)
    
    # 保存文件
    wb.save(output_path)


def translate_label(label: str, language: str) -> str:
    """Translate a label to the specified language.
    
    Uses the advanced TranslationService for better matching and handling of compound labels.
    
    Args:
        label: The English label to translate
        language: Target language ('en' for English, 'zh' for Chinese)
        
    Returns:
        Translated label if found, otherwise original label
    """
    if not isinstance(label, str):
        return label
    
    # Get translation service and use it for translation
    translation_service = get_translation_service()
    return translation_service.translate(label, language)


# Camera brand should remain in its original (English) form; do not translate.


class DirectoryProcessRequest(BaseModel):
    source_path: str = Field(..., description="Directory containing images to process.")
    output_path: str = Field(..., description="Directory to write renamed images into.")
    score_output_path: str = Field(default="", description="Optional path to output score Excel file.")
    recursive: bool = Field(default=True, description="Whether to scan subdirectories.")
    confidence_threshold: float = Field(default=settings.CONFIDENCE_THRESHOLD, ge=0.0, le=1.0)
    # Allow clients to submit any integer >=1 and clamp server-side to the configured maximum.
    max_labels: int = Field(default=settings.MAX_LABELS, ge=1)
    include_camera: bool = Field(default=True, description="Include camera manufacturer in filename.")
    include_type: bool = Field(default=True, description="Include photo type in filename.")
    include_elements: bool = Field(default=True, description="Include AI-recognized elements in filename.")
    element_model: str = Field(default=settings.MODEL_NAME, description="Element recognition model to use.")
    label_language: str = Field(default="en", description="Language for labels: 'en' for English, 'zh' for Chinese.")
    device: str = Field(default="auto", description="Device to use for inference: 'cpu', 'cuda', or 'auto'.")


def _create_job_record(job_id: str, payload: DirectoryProcessRequest, source_dir: str, output_dir: str, total: int) -> dict:
    return {
        "job_id": job_id,
        "status": "queued",
        "message": "Task queued.",
        "source_path": source_dir,
        "output_path": output_dir,
        "model": payload.element_model,
        "validation_model": settings.VALIDATION_MODEL_NAME,
        "confidence_threshold": payload.confidence_threshold,
        "max_labels": payload.max_labels,
        "total": total,
        "processed": 0,
        "progress_percentage": 0,
        "successful": 0,
        "failed": 0,
        "results": [],
        "errors": [],
        "started_at": datetime.utcnow().isoformat() + "Z",
        "completed_at": None,
        "total_duration_seconds": None,
        "total_duration_formatted": None,
        "paused": False,
        "cancelled": False,
    }


def _update_job_progress(job: dict, processed: int, total: int, message: str) -> None:
    job["processed"] = processed
    job["total"] = total
    job["progress_percentage"] = int((processed / total) * 100) if total else 100
    job["message"] = message


def _process_directory_sync(job_id: str, payload: DirectoryProcessRequest, image_paths: list[str], output_dir: str) -> None:
    job = JOB_STORE[job_id]
    job["status"] = "running"
    job["message"] = "Loading recognition models."

    if payload.element_model not in settings.SUPPORTED_ELEMENT_RECOGNITION_MODELS:
        raise HTTPException(
            status_code=400,
            detail=(
                "Unsupported element_model. "
                "Valid values are: "
                f"{', '.join(settings.SUPPORTED_ELEMENT_RECOGNITION_MODELS)}"
            ),
        )

    recognition_service = get_recognition_service(
        payload.element_model,
        settings.VALIDATION_MODEL_NAME,
        payload.device,
        use_fast=True,
    )
    job["device"] = recognition_service.device

    total = len(image_paths)
    results: list[dict] = []
    errors: list[dict] = []

    for index, image_path in enumerate(image_paths):
        # Check if job is paused
        if job.get("paused", False):
            job["status"] = "paused"
            job["message"] = "Task paused by user."
            return

        # Check if job is cancelled
        if job.get("cancelled", False):
            job["status"] = "cancelled"
            job["message"] = "Task cancelled by user."
            return

        original_name = os.path.basename(image_path)
        _update_job_progress(job, index, total, f"Processing {index + 1}/{total}: {original_name}")
        try:
            requested_max_labels = min(payload.max_labels, settings.MAX_LABELS)
            recognition = recognition_service.recognize_image(
                image_path=image_path,
                confidence_threshold=payload.confidence_threshold,
                max_labels=requested_max_labels,
            )
            labels = [item["label"] for item in recognition.get("labels", [])]
            
            # 根据用户选择构建标签列表
            final_labels = []
            language = payload.label_language
            
            # 1. 相机制造商（如果用户选择包含）
            if payload.include_camera:
                camera_make = get_camera_make(image_path)
                if camera_make:
                    # 保持相机品牌始终为原始英文表示（不随语言切换）
                    final_labels.append(camera_make)
                else:
                    final_labels.append("unknown")
            
            # 2. 照片类型（如果用户选择包含）- 使用CLIP分类器的14大类结果
            if payload.include_type:
                category_en = recognition.get("category_en")
                category_zh = recognition.get("category_zh")
                if category_en and category_zh:
                    # 根据语言选择标签
                    photo_type = category_zh if language == "zh" else category_en
                    # 将照片类型转换为文件名安全的格式
                    sanitized_type = "_".join(photo_type.lower().split())
                    final_labels.append(sanitized_type)
            
            # 3. AI识别的元素（如果用户选择包含）- 使用选定的语言
            if payload.include_elements:
                for label in labels:
                    translated_label = translate_label(label, language)
                    final_labels.append(translated_label)
            
            # 使用构建的标签列表
            labels = final_labels

            # 获取照片类别用于文件夹分类（如果用户选择了包含类型）
            category_folder_name = None
            if payload.include_type:
                category_en = recognition.get("category_en")
                category_zh = recognition.get("category_zh")
                if category_en and category_zh:
                    category_folder_name = category_zh if language == "zh" else category_en

            if not labels:
                # 没有有效标签时，不进入类别文件夹，直接保留原文件名
                output_path = build_unique_path_for_name(output_dir, original_name)
                copy_to_output(image_path, output_path)
                results.append(
                    {
                        "index": index,
                        "source_path": image_path,
                        "output_path": output_path,
                        "original_filename": original_name,
                        "renamed_filename": os.path.basename(output_path),
                        "labels": [],
                        "status": "kept_original_name",
                        "message": "No elements met the confidence threshold, so the original filename was kept.",
                    }
                )
            else:
                extension = get_file_extension(original_name) or "jpg"
                # 如果有类别信息且用户选择了包含类型，则按类别分文件夹
                if category_folder_name:
                    output_path = build_category_output_path(output_dir, category_folder_name, labels, extension)
                else:
                    output_path = build_unique_output_path(output_dir, labels, extension)
                copy_to_output(image_path, output_path)
                # 翻译并构造返回给前端的标签列表（用于 UI 显示）
                display_labels = []
                for item in recognition.get("labels", []):
                    lab = item.get("label", "")
                    display = translate_label(lab, language)
                    display_labels.append(
                        {
                            "label": display,
                            "confidence_percentage": item.get("confidence_percentage", round(item.get("score", 0) * 100, 2)),
                        }
                    )

                results.append(
                    {
                        "index": index,
                        "source_path": image_path,
                        "output_path": output_path,
                        "original_filename": original_name,
                        "renamed_filename": os.path.basename(output_path),
                        "labels": display_labels,
                        "status": "renamed",
                    }
                )
        except Exception as exc:  # pragma: no cover - defensive error surface
            logger.exception("Failed to process %s", image_path)
            errors.append(
                {
                    "index": index,
                    "source_path": image_path,
                    "filename": original_name,
                    "error": str(exc),
                }
            )

        job["results"] = results
        job["errors"] = errors
        job["successful"] = len(results)
        job["failed"] = len(errors)
        _update_job_progress(job, index + 1, total, f"Processed {index + 1}/{total}: {original_name}")

    job["status"] = "completed"
    job["completed_at"] = datetime.utcnow().isoformat() + "Z"
    job["message"] = "Processing completed."
    job["progress_percentage"] = 100

    # 计算总处理时长
    try:
        from dateutil import parser
        started = parser.isoparse(job["started_at"])
        completed = parser.isoparse(job["completed_at"])
        total_seconds = (completed - started).total_seconds()
        job["total_duration_seconds"] = round(total_seconds, 2)
        job["total_duration_formatted"] = _format_duration(total_seconds)
        logger.info(f"Job {job_id} completed in {job['total_duration_formatted']}")
    except Exception as e:
        logger.warning(f"Could not calculate processing duration: {e}")
        # 使用一个估算值作为后备
        job["total_duration_seconds"] = 0.0
        job["total_duration_formatted"] = "< 1 秒"

    # 生成照片评分Excel表格
    if payload.score_output_path:
        try:
            _generate_score_excel(results, errors, payload.score_output_path)
            logger.info(f"Score Excel file generated at: {payload.score_output_path}")
            job["score_output_path"] = payload.score_output_path
        except Exception as e:
            logger.exception(f"Failed to generate score Excel: {e}")


async def _run_directory_job(job_id: str, payload: DirectoryProcessRequest, image_paths: list[str], output_dir: str) -> None:
    try:
        await asyncio.to_thread(_process_directory_sync, job_id, payload, image_paths, output_dir)
    except Exception as exc:  # pragma: no cover - defensive error surface
        logger.exception("Directory job %s failed", job_id)
        job = JOB_STORE[job_id]
        job["status"] = "failed"
        job["message"] = str(exc)
        job["completed_at"] = datetime.utcnow().isoformat() + "Z"


@router.post("/process-directory")
async def process_directory(payload: DirectoryProcessRequest) -> dict:
    """Create a background directory processing job."""
    # Clamp client-provided values to server configuration to avoid 422 validation errors
    try:
        payload.max_labels = int(min(payload.max_labels, settings.MAX_LABELS))
        source_dir = os.path.abspath(os.path.expanduser(payload.source_path.strip()))
        output_dir = ensure_directory(payload.output_path)
        image_paths = iter_image_files(source_dir, recursive=payload.recursive)
    except (FileNotFoundError, NotADirectoryError) as exc:
        raise HTTPException(status_code=400, detail=str(exc)) from exc
    except OSError as exc:
        raise HTTPException(status_code=400, detail=f"Invalid path configuration: {exc}") from exc

    if payload.element_model not in settings.SUPPORTED_ELEMENT_RECOGNITION_MODELS:
        raise HTTPException(
            status_code=400,
            detail=(
                "Unsupported element_model. "
                "Valid values are: "
                f"{', '.join(settings.SUPPORTED_ELEMENT_RECOGNITION_MODELS)}"
            ),
        )

    if not image_paths:
        raise HTTPException(status_code=400, detail="No supported image files were found in the source directory.")

    job_id = uuid4().hex
    JOB_STORE[job_id] = _create_job_record(job_id, payload, source_dir, output_dir, len(image_paths))
    asyncio.create_task(_run_directory_job(job_id, payload, image_paths, output_dir))

    return {
        "job_id": job_id,
        "status": "queued",
        "total": len(image_paths),
        "progress_percentage": 0,
        "message": "Task queued.",
    }


@router.get("/process-directory/{job_id}")
async def get_directory_job(job_id: str) -> dict:
    """Fetch the current status of a directory processing job."""
    job = JOB_STORE.get(job_id)
    if not job:
        raise HTTPException(status_code=404, detail="Job not found.")
    return job


@router.post("/process-directory/{job_id}/pause")
async def pause_directory_job(job_id: str) -> dict:
    """Pause a running directory processing job."""
    job = JOB_STORE.get(job_id)
    if not job:
        raise HTTPException(status_code=404, detail="Job not found.")

    if job["status"] not in ["running", "queued"]:
        raise HTTPException(status_code=400, detail="Job is not in a pausable state.")

    job["paused"] = True
    job["status"] = "paused"
    job["message"] = "Task paused by user."

    return {"job_id": job_id, "status": "paused", "message": "Task paused successfully."}


@router.post("/process-directory/{job_id}/cancel")
async def cancel_directory_job(job_id: str) -> dict:
    """Cancel a running directory processing job."""
    job = JOB_STORE.get(job_id)
    if not job:
        raise HTTPException(status_code=404, detail="Job not found.")

    if job["status"] in ["completed", "failed", "cancelled"]:
        raise HTTPException(status_code=400, detail=f"Job is already {job['status']}.")

    job["cancelled"] = True
    job["status"] = "cancelled"
    job["message"] = "Task cancelled by user."

    return {"job_id": job_id, "status": "cancelled", "message": "Task cancelled successfully."}


@router.post("/process-directory/{job_id}/resume")
async def resume_directory_job(job_id: str) -> dict:
    """Resume a paused directory processing job."""
    job = JOB_STORE.get(job_id)
    if not job:
        raise HTTPException(status_code=404, detail="Job not found.")

    if job["status"] != "paused":
        raise HTTPException(status_code=400, detail="Job is not paused.")

    job["paused"] = False
    job["status"] = "running"
    job["message"] = "Task resumed."

    # Create a new task to continue processing from where it left off
    # We need to reconstruct the image paths and continue from the current index
    try:
        from ..utils.file_handler import iter_image_files
        image_paths = list(iter_image_files(job["source_path"], recursive=True))
        remaining_paths = image_paths[job["processed"]:]

        if remaining_paths:
            # Create a mock payload for resuming
            resume_payload = DirectoryProcessRequest(
                source_path=job["source_path"],
                output_path=job["output_path"],
                recursive=True,
                confidence_threshold=job["confidence_threshold"],
                max_labels=job["max_labels"],
                include_camera=True,
                include_type=True,
                include_elements=True,
                element_model=job["model"],
                label_language="en",
            )
            asyncio.create_task(_run_directory_job(job_id, resume_payload, remaining_paths, job["output_path"]))
        else:
            # No more files to process
            job["status"] = "completed"
            job["message"] = "Processing completed."
            job["completed_at"] = datetime.utcnow().isoformat() + "Z"
            
            # 计算总处理时长（即使是从暂停恢复的）
            try:
                from dateutil import parser
                started = parser.isoparse(job["started_at"])
                completed = parser.isoparse(job["completed_at"])
                total_seconds = (completed - started).total_seconds()
                job["total_duration_seconds"] = round(total_seconds, 2)
                job["total_duration_formatted"] = _format_duration(total_seconds)
            except Exception as e:
                logger.warning("Could not calculate processing duration: %s", e)
                job["total_duration_seconds"] = None
                job["total_duration_formatted"] = None

    except Exception as exc:
        logger.exception("Failed to resume job %s", job_id)
        job["status"] = "failed"
        job["message"] = f"Failed to resume: {str(exc)}"

    return {"job_id": job_id, "status": job["status"], "message": job["message"]}


@router.get("/health")
async def health_check() -> dict:
    """Service health information."""
    return {
        "status": "healthy",
        "model": settings.MODEL_NAME,
        "validation_model": settings.VALIDATION_MODEL_NAME,
        "confidence_threshold": settings.CONFIDENCE_THRESHOLD,
        "max_labels": settings.MAX_LABELS,
        "device_preference": settings.DEVICE_PREFERENCE,
    }


@router.get("/info")
async def get_info() -> dict:
    """Frontend configuration payload."""
    # 不要在这里加载模型！否则会阻塞 /info 接口响应
    return {
        "model": settings.MODEL_NAME,
        "validation_model": settings.VALIDATION_MODEL_NAME,
        "element_model": settings.MODEL_NAME,
        "element_models": [
            {
                "name": name,
                "label": (value["label"] if isinstance(value, dict) else value),
                "desc": (value.get("desc") if isinstance(value, dict) else None),
            }
            for name, value in settings.SUPPORTED_ELEMENT_RECOGNITION_MODELS.items()
        ],
        "model_source": settings.MODEL_NAME,
        "validation_model_source": settings.VALIDATION_MODEL_NAME,
        "confidence_threshold": settings.CONFIDENCE_THRESHOLD,
        "max_labels": settings.MAX_LABELS,
        "device_preference": settings.DEVICE_PREFERENCE,
        "device": settings.DEVICE_PREFERENCE,
        "cuda_available": torch.cuda.is_available(),
        "torch_version": torch.__version__,
        "allowed_extensions": sorted(settings.ALLOWED_EXTENSIONS),
    }
